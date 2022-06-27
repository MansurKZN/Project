import numbers
import os
from datetime import date
import datetime

import openpyxl

from django.contrib.auth import logout
from django.contrib.auth.views import LoginView
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views import View
from openpyxl.styles import Border, Side, Font

from app.forms import LoginForm, PasswordChangeForm
from app.models import Project, Engineer, WorkType, WorkReport, Manager
from project import settings


class Login(LoginView):
    form_class = LoginForm
    template_name = 'login.html'

    def get_success_url(self):
        user = self.request.user
        if user.is_superuser:
            return reverse_lazy('admin')
        else:
            return reverse_lazy('main')

def logout_user(request):
    logout(request)
    return redirect('login')


class MainPage(View):

    def get(self, request):
        user = request.user
        projects = list(map(lambda x: x.name, Project.objects.all()))
        engineers = list(map(lambda x: x.full_name, Engineer.objects.all()))
        work_types = list(map(lambda x: x.name, WorkType.objects.all()))
        arr = []
        work_reports = WorkReport.objects.filter(manager=user)
        for report in work_reports:
            a = []
            a.append(report.id)
            a.append(report.date.__format__("%Y-%m-%d %H:%M"))
            a.append(report.project.name)
            a.append(report.engineer.full_name)
            a.append(report.work_type.name)
            a.append(report.period)
            a.append(report.text)
            arr.append(a)
        date = datetime.datetime.now().__format__("%Y-%m-%dT%H:%M")

        paginator = Paginator(arr, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, 'main.html', {"user": user, 'projects': projects, 'engineers': engineers, 'work_types': work_types, 'date': date, 'arr': arr, 'page_obj': page_obj})

    def post(self, request):
        if request.POST.get("update_id"):
            work_report = WorkReport.objects.get(id=request.POST.get("update_id"))
            work_report.date = request.POST.get("update_date")
            work_report.project = Project.objects.filter(name=request.POST.get("update_project"))[0]
            work_report.engineer = Engineer.objects.filter(full_name=request.POST.get("update_engineer"))[0]
            work_report.work_type = WorkType.objects.filter(name=request.POST.get("update_work_type"))[0]
            work_report.period = request.POST.get("update_period")
            work_report.text = request.POST.get("update_text")
            work_report.save()
        elif request.POST.get("delete"):
            WorkReport.objects.get(id=request.POST.get("delete")).delete()
        else:
            user = request.user
            project = Project.objects.filter(name=request.POST.get("project"))[0]
            engineer = Engineer.objects.filter(full_name=request.POST.get("engineer"))[0]
            work_type = WorkType.objects.filter(name=request.POST.get("work_type"))[0]
            work_report = WorkReport(manager=user, date=request.POST.get("date"), project=project, engineer=engineer, work_type=work_type, period=request.POST.get("period"), text=request.POST.get("text"))
            work_report.save()
        return redirect('/main')


class Profie(View):

    def get(self, request):
        user = request.user
        form = PasswordChangeForm()
        return render(request, 'profile.html', {"user": user, 'form': form})

    def post(self, request):
        form = PasswordChangeForm(request.POST)
        user = request.user
        message = ''
        error = ''
        if user.check_password(form.data.get('password_old')):
            if form.data.get('password_new') == form.data.get('password_retype'):
                user.set_password(form.data.get('password_new'))
                user.save()
                message = 'Password changed successfully'
            else:
                error = 'New password and retype password not equal'
                form.add_error(None, 'Password change error')
        else:
            error = 'Old password is incorrect'
            form.add_error(None, 'Password change error')
        return render(request, 'profile.html', {"user": user, 'form': form, 'message': message, 'error': error})


class AdminInfo(View):

    def get(self, request):
        user = request.user
        arr = []
        date_start = date(date.today().year, date.today().month, 1)
        date_end = date(date.today().year, date.today().month + 1, 1)
        projects = Project.objects.all()
        columns = []
        columns.append("")
        columns.append("")
        columns.append("Раб часы")
        for project in projects:
            columns.append(project.name)

        for manager in Manager.objects.filter(is_superuser=False):
            a = []
            work_time = manager.work_time
            a.append(manager.work_position)
            a.append(f'{manager.first_name} {manager.last_name}')
            a.append(work_time)
            for project in projects:
                time_for_project = WorkReport.objects.filter(manager=manager, date__range=(date_start, date_end), project=project).aggregate(Sum('period')).get('period__sum')
                if time_for_project:
                    percent_project_work = round(((int(time_for_project) / 60) * 100 / work_time) / 100, 2)
                else:
                    percent_project_work = 0.0
                a.append(percent_project_work)
            arr.append(a)
        return render(request, 'admin.html', {'user': user.username, 'columns': columns, 'arr': arr, 'date': datetime.datetime.now().strftime('%Y-%m')})

    def post(self, request):
        for item in os.listdir(f"{settings.BASE_DIR}/static/"):
            if item.endswith(".xlsx"):
                os.remove(f"{settings.BASE_DIR}/static/{item}")

        arr = []
        if request.POST.get("date"):
            d = datetime.datetime.strptime(request.POST.get("date"), '%Y-%m')
            date_start = date(d.year, d.month, 1)
            date_end = date(date_start.year, date_start.month + 1, 1)
        else:
            d = datetime.datetime.strptime(request.POST.get("export_date"), '%Y-%m')
            date_start = date(d.year, d.month, 1)
            date_end = date(date_start.year, date_start.month + 1, 1)

        projects = Project.objects.all()
        columns = []
        columns.append("")
        columns.append("")
        columns.append("Раб часы")
        max_name_len = 0
        max_proj_len = 0
        for project in projects:
            if len(project.name) > max_proj_len:
                max_proj_len = len(project.name)
            columns.append(project.name)
        arr.append(columns)
        for manager in Manager.objects.filter(is_superuser=False):
            if len(f'{manager.first_name} {manager.last_name}') > max_name_len:
                max_name_len = len(f'{manager.first_name} {manager.last_name}')
            a = []
            work_time = manager.work_time
            a.append(manager.work_position)
            a.append(f'{manager.first_name} {manager.last_name}')
            a.append(work_time)
            for project in projects:
                time_for_project = WorkReport.objects.filter(manager=manager, date__range=(date_start, date_end),
                                                             project=project).aggregate(Sum('period')).get(
                    'period__sum')
                if time_for_project:
                    percent_project_work = round(((int(time_for_project) / 60) * 100 / work_time) / 100, 2)
                else:
                    percent_project_work = 0.0
                a.append(percent_project_work)
            arr.append(a)

        if request.POST.get("date"):
            columns = arr[0]
            arr.pop(0)
            return render(request, 'admin.html', {'user': request.user.username, 'columns': columns, 'arr': arr, 'date': date_start.strftime('%Y-%m')})


        wb = openpyxl.Workbook()
        sheet = wb.active
        for row in arr:

            sheet.append(row)

        for i in sheet.columns:
            sheet.column_dimensions[i[0].column_letter].width = max_proj_len + 3

        sheet.column_dimensions['B'].width = max_name_len + 1

        last_row_num = 0
        last_column_char = ''
        for column in sheet.columns:
            for elem in column:
                if isinstance(elem.value, numbers.Number):
                    sheet.cell(column=elem.column, row=elem.row).border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='00000000'),
                                         top=Side(border_style='thin', color='00000000'),
                                         bottom=Side(border_style='thin', color='00000000'))
                last_row_num = elem.row

            if column[0].column > 3:
                sheet.cell(row=last_row_num+1, column=column[0].column).value = f'=SUM({column[0].column_letter}2:{column[0].column_letter}{last_row_num})'
                sheet.cell(row=last_row_num+1, column=column[0].column).font = Font(bold=True)
            last_column_char = column[0].column_letter

        for row in sheet.rows:
            if row[0].row > 1 and row[0].row <= last_row_num:
                sheet.cell(row=row[0].row, column=len(row)+1).value = f'=SUM(D{row[0].row}:{last_column_char}{row[0].row})'
                sheet.cell(row=row[0].row, column=len(row)+1).font = Font(bold=True)

        file_name = f'Work_time_{date_start}_{date_end}.xlsx'
        wb.save(f'{settings.BASE_DIR}/static/{file_name}')

        return redirect(f'/static/{file_name}')


class TimeControl(View):

    def get(self, request):
        user = request.user
        arr = []
        sum = 0
        date_start = date.today()
        date_end = date_start + datetime.timedelta(days=1)
        projects = []
        for i in WorkReport.objects.filter(manager=user, date__range=(date_start, date_end)):
            projects.append(i.project)

        for project in list(set(projects)):
            a = []
            a.append(project.name)
            company_sum = WorkReport.objects.filter(manager=user, date__range=(date_start, date_end), project=project).aggregate(Sum('period')).get('period__sum')
            sum += company_sum
            a.append(company_sum)
            arr.append(a)
        a = ['', sum]
        arr.append(a)
        return render(request, 'timecontrol.html', {'user': user, 'arr': arr})

    def post(self, request):
        user = request.user
        arr = []
        sum = 0
        date_start = datetime.date.fromisoformat(request.POST.get("date"))
        date_end = date_start + datetime.timedelta(days=1)
        projects = []
        for i in WorkReport.objects.filter(manager=user, date__range=(date_start, date_end)):
            projects.append(i.project)
        for project in list(set(projects)):
            a = []
            a.append(project.name)
            company_sum = WorkReport.objects.filter(manager=user, date__range=(date_start, date_end),
                                                    project=project).aggregate(Sum('period')).get('period__sum')
            sum += company_sum
            a.append(company_sum)
            arr.append(a)
        a = ['', sum]
        arr.append(a)
        return render(request, 'timecontrol.html', {'user': user, 'arr': arr})